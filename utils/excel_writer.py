# utils/excel_writer.py

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, timezone
from pathlib import Path
from typing import List, Dict
import os


# SahaKayitlari sheet column headers
SAHA_KAYITLARI_HEADERS = [
    'SahaID',
    'PONumber',
    'HizmetKodu',
    'OrderQty',
    'Price',
    'NetValue',
    'SourceFileName',
    'ImportedAt'
]


def ensure_master_workbook(master_path: str) -> Workbook:
    """
    Creates or opens Master Excel workbook.
    
    Args:
        master_path: Path to Master.xlsx file
    
    Returns:
        Workbook: openpyxl Workbook object
    """
    master_file = Path(master_path)
    
    # Create parent directory if not exists
    master_file.parent.mkdir(parents=True, exist_ok=True)
    
    if master_file.exists():
        # Load existing workbook
        try:
            wb = load_workbook(master_path)
            return wb
        except Exception as e:
            raise ValueError(f"Cannot open existing Master.xlsx: {str(e)}")
    else:
        # Create new workbook
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create Katalog sheet
        wb.create_sheet('Katalog')
        
        # Create SahaKayitlari sheet
        wb.create_sheet('SahaKayitlari')
        
        # Save
        wb.save(master_path)
        
        return wb


def ensure_sheet_and_headers(workbook: Workbook, sheet_name: str = "SahaKayitlari", 
                             headers: List[str] = None) -> None:
    """
    Ensures sheet exists and has proper headers.
    
    Args:
        workbook: openpyxl Workbook
        sheet_name: Name of the sheet
        headers: List of column headers
    """
    if headers is None:
        headers = SAHA_KAYITLARI_HEADERS
    
    # Create sheet if not exists
    if sheet_name not in workbook.sheetnames:
        workbook.create_sheet(sheet_name)
    
    sheet = workbook[sheet_name]
    
    # Check if headers exist (first row)
    if sheet.max_row == 1 or all(cell.value is None for cell in sheet[1]):
        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_idx, value=header)


def append_saha_kayitlari_rows(master_path: str, rows: List[Dict]) -> int:
    """
    Appends rows to SahaKayitlari sheet in Master Excel.
    
    Args:
        master_path: Path to Master.xlsx
        rows: List of dictionaries with keys matching SAHA_KAYITLARI_HEADERS
    
    Returns:
        int: Number of rows appended
    
    Raises:
        ValueError: If Excel operations fail
    """
    try:
        # Open workbook
        wb = ensure_master_workbook(master_path)
        
        # Ensure sheet and headers
        ensure_sheet_and_headers(wb, "SahaKayitlari", SAHA_KAYITLARI_HEADERS)
        
        sheet = wb["SahaKayitlari"]
        
        # Get next row
        next_row = sheet.max_row + 1
        rows_added = 0
        
        # Append rows
        for row_data in rows:
            # Add ImportedAt timestamp if not present
            if 'ImportedAt' not in row_data or row_data['ImportedAt'] is None:
                row_data['ImportedAt'] = datetime.now(timezone.utc).isoformat()
            
            # Write data in column order
            for col_idx, header in enumerate(SAHA_KAYITLARI_HEADERS, start=1):
                value = row_data.get(header, '')
                sheet.cell(row=next_row, column=col_idx, value=value)
            
            next_row += 1
            rows_added += 1
        
        # Save workbook
        wb.save(master_path)
        wb.close()
        
        return rows_added
    
    except Exception as e:
        raise ValueError(f"Error appending rows to SahaKayitlari: {str(e)}")


def get_existing_records(master_path: str) -> List[Dict]:
    """
    Retrieves all existing records from SahaKayitlari sheet.
    
    Args:
        master_path: Path to Master.xlsx
    
    Returns:
        List[Dict]: List of existing records
    """
    try:
        if not Path(master_path).exists():
            return []
        
        wb = load_workbook(master_path, read_only=True)
        
        if "SahaKayitlari" not in wb.sheetnames:
            wb.close()
            return []
        
        sheet = wb["SahaKayitlari"]
        
        # Get headers from first row
        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(cell.value)
        
        if not headers:
            wb.close()
            return []
        
        # Read all rows
        records = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row):
                record = {}
                for idx, header in enumerate(headers):
                    if idx < len(row):
                        record[header] = row[idx]
                records.append(record)
        
        wb.close()
        return records
    
    except Exception as e:
        raise ValueError(f"Error reading existing records: {str(e)}")
